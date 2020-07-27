import smtplib
import ssl
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
import pandas as pd
from flask import Flask, request, render_template

app = Flask(__name__)
cols = ['Student ID', 'Full Name', 'Email']


@app.route("/upload", methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':

        f = request.files['file']
        data_xls = pd.read_excel(f, usecols=cols)

        # try to open an existing workbook
        try:
            writer = openpyxl.load_workbook('demo.xlsx')
        except FileNotFoundError:
            print("The workbook does not exist. Creating one...")
            writer = pd.ExcelWriter('demo.xlsx', engine='xlsxwriter', )
            writer.save()
            writer = openpyxl.load_workbook('demo.xlsx')
            sheet = writer.worksheets[0]
            new_row_data = [['Student ID']]
            for row_data in new_row_data:
                # Append Row Values
                sheet.append(row_data)
            writer.save('demo.xlsx')
            writer = openpyxl.load_workbook('demo.xlsx')

        # read existing file
        reader = pd.read_excel(r'demo.xlsx')

        process_data(data_xls, reader, writer)
    return render_template('index.html')


def process_data(data_excel, workbook, writer):
    sheet = writer.worksheets[0]
    for index, row in data_excel.iterrows():
        workbook = pd.read_excel(r'demo.xlsx')
        student_id = ""

        if pd.notna(row["Student ID"]):
            student_id = row["Student ID"]
        else:
            continue

        if row["Student ID"] not in workbook["Student ID"].tolist():
            if pd.notna(row["Full Name"]):
                full_name = row["Full Name"]
                student_full_name = full_name.split(' ')
                student_first_name = student_full_name[0]
                student_last_name = student_full_name[1]

            print(full_name + ',' + student_first_name + ',' + student_last_name)
            send_email(row["Email"])
            new_row_data = [[student_id]]
            for row_data in new_row_data:
                # Append Row Values
                sheet.append(row_data)

            writer.save('demo.xlsx')
    writer.close()



@app.route("/", methods=['GET'])
def export_records():
    return render_template('index.html')


def send_email(receiver_email):
    sender_email = "sender_emailid"
    password = "password"

    message = MIMEMultipart("alternative")
    message["Subject"] = "Welcome to Medicaps Family"
    message["From"] = sender_email
    message["To"] = receiver_email

    html = """\
        <html>
          <body>
            <p>Hi,<br><br>
               Congratulations! You are enrolled in Medicaps University.<br>
               You are about to begin one of the most exciting times in your life.<br>
               Your session will start on 20-July-2020.<br>
               For further details about the University, visit <a href ="http://www.medicaps.ac.in/">Medicaps University</a><br>
               For transport facility concern, visit <a href ="transport@medicapsuniversity.com">Transport</a><br>
               For books and study material, visit <a href ="studymaterial@medicapsuniveristy.com">Study Material</a><br><br>
               Thanks,<br>
               Medicaps Team

            <br>
            </p>
          </body>
        </html>
        """

    # Turn these into plain/html MIMEText objects
    part = MIMEText(html, "html")

    # Add HTML/plain-text parts to MIMEMultipart message
    # The email client will try to render the last part first

    message.attach(part)

    # Create secure connection with server and send email
    # context = ssl.create_default_context()
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.ehlo()
    server.starttls()
    server.login(sender_email, password)
    server.sendmail(sender_email, receiver_email, message.as_string())
    server.close()


if __name__ == "__main__":
    app.run(debug=True)
